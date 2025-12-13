#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Enhanced Symbol Filter: فیلتر پیشرفته با معیارهای دقیق‌تر
"""

import pandas as pd
import logging
import os
import json
from datetime import date
import argparse
from typing import Optional

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class EnhancedSymbolFilter:
    """
    فیلتر نمادهای معتبر برای تحلیل تکنیکال
    بر اساس:
    - نوع نماد (سهام عادی فقط)
    - فعالیت (حجم معاملات)
    - معیارهای اقتصادی (P/E, EPS)
    """
    
    # نمادهای غیرمعتبر
    INVALID_KEYWORDS = {
        'warrant_rights': ['د', 'ـر', 'ـح', 'حق'],
        'funds_etf': ['صندوق', 'ETF', 'شاخصی'],
        'bonds': ['صکوك', 'اوراق', 'درآمد ثابت', 'بدهی'],
        'murabaha': ['مرابحه'],
        'housing': ['مسکن', 'تس'],
        'forex_crypto': ['ارز', 'تالار', 'کریپتو'],
        'inactive': ['حذف', 'تعلیق', 'معلق', 'غیرفعال'],
        'government': ['دولتی', 'خزانه', 'اسلامی'],
        'real_estate': ['املاک', 'مستغلات', 'سرمایه‌گذاری'],
        'investment': ['ص.س.', 'سرمایه گذاری', 'سرمایه‌گذاری'],
    }
    
    # بخش‌های صنعتی معتبر
    VALID_SECTORS = [
        'سیمان', 'فولاد', 'خودرو', 'پتروشیمی', 'نفت', 'گاز',
        'بانک', 'بیمه', 'مواد', 'معادن', 'الومینیوم', 'مس',
        'انرژی', 'برق', 'آب', 'مخابرات', 'دارو', 'غذا',
        'شیر', 'کشاورزی', 'خدمات', 'گردشگری', 'حمل و نقل',
        'صنعتی', 'عمومی', 'شیمی', 'ریخته‌گری', 'نساجی',
        'سرامیک', 'کاشی', 'شیشه', 'چینی', 'سازه‌ای'
    ]
    
    def __init__(self):
        self.filters_applied = {}
        self.removed_symbols = []
    
    def check_invalid_keywords(self, symbol: str, name: str) -> tuple[bool, str]:
        """بررسی کلمات کلیدی نامعتبر"""
        symbol_full = f"{symbol} {name}".upper()
        
        for category, keywords in self.INVALID_KEYWORDS.items():
            for keyword in keywords:
                if keyword in symbol_full:
                    return False, f"[{category}] {keyword}"
        
        return True, ""
    
    def check_valid_sector(self, name: str) -> bool:
        """بررسی آیا بخش صنعتی معتبر است"""
        name_upper = name.upper()
        
        # بررسی اینکه نام شامل حداقل یک بخش معتبر است
        for sector in self.VALID_SECTORS:
            if sector in name_upper:
                return True
        
        return False
    
    def check_numeric_validity(self, row: pd.Series) -> tuple[bool, str]:
        """بررسی داده‌های عددی"""
        
        # بررسی حجم معاملات
        try:
            volume = pd.to_numeric(row.get('حجم', 0), errors='coerce')
            if pd.isna(volume) or volume <= 0:
                return False, "حجم صفر یا نامعتبر"
        except:
            pass
        
        # بررسی قیمت آخرین معامله
        try:
            last_price = pd.to_numeric(row.get('آخرین معامله - مقدار', 0), errors='coerce')
            if pd.isna(last_price) or last_price <= 100:
                return False, "قیمت بسیار پایین یا نامعتبر"
        except:
            pass
        
        return True, ""
    
    def filter_symbols(self, df: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
        """فیلتر کردن نمادهای معتبر"""
        
        valid_rows = []
        
        for idx, row in df.iterrows():
            symbol = str(row.get('نماد', '')).strip()
            name = str(row.get('نام', '')).strip()
            
            # فیلتر 1: کلمات کلیدی نامعتبر
            is_valid, reason = self.check_invalid_keywords(symbol, name)
            if not is_valid:
                self._record_removal(symbol, name, reason)
                continue
            
            # فیلتر 2: بخش صنعتی معتبر
            if not self.check_valid_sector(name):
                self._record_removal(symbol, name, "[sector] نوع صنعت نامعتبر")
                continue
            
            # فیلتر 3: داده‌های عددی
            is_valid, reason = self.check_numeric_validity(row)
            if not is_valid:
                self._record_removal(symbol, name, reason)
                continue
            
            valid_rows.append(idx)
        
        # برگردانید موارد معتبر
        result = df.loc[valid_rows].reset_index(drop=True)
        
        # مرتب‌سازی بر اساس حجم معاملات (اگر موجود باشد)
        if 'حجم' in result.columns:
            try:
                result['حجم_numeric'] = pd.to_numeric(result['حجم'], errors='coerce')
                result = result.sort_values('حجم_numeric', ascending=False, na_position='last')
                result = result.drop('حجم_numeric', axis=1)
            except:
                pass
        
        # برگردانید برترین N نماد
        return result.head(top_n).reset_index(drop=True)
    
    def _record_removal(self, symbol: str, name: str, reason: str):
        """ثبت نمادهای حذف‌شده"""
        self.removed_symbols.append({
            'symbol': symbol,
            'name': name,
            'reason': reason
        })
        
        if reason not in self.filters_applied:
            self.filters_applied[reason] = 0
        self.filters_applied[reason] += 1
    
    def print_report(self):
        """چاپ گزارش فیلتر"""
        total_removed = len(self.removed_symbols)
        
        logger.info("\n" + "="*80)
        logger.info("📊 گزارش فیلتر شدگی")
        logger.info("="*80)
        logger.info(f"❌ کل نمادهای حذف شده: {total_removed}")
        logger.info("\n📋 دلایل حذف:")
        
        for reason, count in sorted(self.filters_applied.items(), 
                                     key=lambda x: x[1], reverse=True):
            logger.info(f"  - {reason}: {count}")


def main():
    """برنامه اصلی"""
    
    parser = argparse.ArgumentParser(description='Enhanced symbol filter outputs')
    parser.add_argument('--project-name', default='J.M_Stock_Market', help='Project name to include in outputs')
    parser.add_argument('--contact-email', default=None, help='Contact email to embed in outputs')
    parser.add_argument('--top-n', type=int, default=20, help='Number of top symbols to keep')
    parser.add_argument('--save-pdf', action='store_true', help='Also export results to PDF if reportlab is available')
    parser.add_argument('--save-excel', action='store_true', help='Also export results to Excel')
    args = parser.parse_args()

    # if user didn't specify flags, produce both Excel and PDF by default
    if not args.save_excel and not args.save_pdf:
        args.save_excel = True
        args.save_pdf = True

    print("\n" + "="*80)
    print(f"🎯 فیلتر پیشرفتۀ نمادها برای تحلیل تکنیکال ({args.project_name})")
    print("="*80 + "\n")
    
    # بارگذاری داده‌ها
    logger.info("📥 بارگذاری فایل CSV...")
    
    try:
        df = pd.read_csv('data/indexes/symbols.csv', skiprows=2, encoding='utf-8')
        df.columns = df.columns.str.strip()
        
        logger.info(f"✅ بارگذاری موفق: {len(df)} نماد")
        
    except Exception as e:
        logger.error(f"❌ خطا: {e}")
        return
    
    # فیلتر کردن
    logger.info("\n🔍 اعمال فیلترها...")
    
    filter_obj = EnhancedSymbolFilter()
    top_20 = filter_obj.filter_symbols(df, top_n=args.top_n)
    
    # گزارش
    filter_obj.print_report()
    
    # ذخیره‌ی فایل
    logger.info("\n💾 ذخیره‌ی فایل...")
    # ایجاد پوشۀ خروجی با تاریخ امروز
    out_dir = os.path.join('outputs', date.today().isoformat())
    os.makedirs(out_dir, exist_ok=True)

    # نام فایلها شامل نام پروژه و تاریخ
    safe_proj = args.project_name.replace(' ', '_')
    csv_name = f"top_symbols_{safe_proj}_{date.today().isoformat()}.csv"
    top_path = os.path.join(out_dir, csv_name)
    top_20.to_csv(top_path, index=False, encoding='utf-8')
    logger.info(f"✅ ذخیره شد: {top_path}\n")

    # ذخیره‌ی نمادهای حذف‌شده و خلاصهٔ فیلترها
    try:
        if filter_obj.removed_symbols:
            removed_df = pd.DataFrame(filter_obj.removed_symbols)
            removed_name = f"removed_symbols_{safe_proj}_{date.today().isoformat()}.csv"
            removed_path = os.path.join(out_dir, removed_name)
            removed_df.to_csv(removed_path, index=False, encoding='utf-8')
            logger.info(f"✅ ذخیره شد: {removed_path}")

        # ذخیره‌ی خلاصهٔ فیلترها به صورت JSON
        filters_name = f"filters_summary_{safe_proj}_{date.today().isoformat()}.json"
        filters_path = os.path.join(out_dir, filters_name)
        with open(filters_path, 'w', encoding='utf-8') as fh:
            json.dump(filter_obj.filters_applied, fh, ensure_ascii=False, indent=2)
        logger.info(f"✅ ذخیره شد: {filters_path}")

        # ذخیرهٔ متادیتا شامل نام پروژه و ایمیل تماس
        metadata = {
            'project_name': args.project_name,
            'contact_email': args.contact_email,
            'date': date.today().isoformat(),
            'row_count': len(top_20)
        }
        metadata_path = os.path.join(out_dir, f"metadata_{safe_proj}_{date.today().isoformat()}.json")
        with open(metadata_path, 'w', encoding='utf-8') as fh:
            json.dump(metadata, fh, ensure_ascii=False, indent=2)
        logger.info(f"✅ ذخیره شد: {metadata_path}")

        # ذخیره در Excel در صورت درخواست
        if args.save_excel:
            try:
                excel_name = f"top_symbols_{safe_proj}_{date.today().isoformat()}.xlsx"
                excel_path = os.path.join(out_dir, excel_name)
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    top_20.to_excel(writer, sheet_name='TopSymbols', index=False)
                    if filter_obj.removed_symbols:
                        removed_df.to_excel(writer, sheet_name='Removed', index=False)
                    # metadata sheet
                    meta_df = pd.DataFrame(list(metadata.items()), columns=['key', 'value'])
                    meta_df.to_excel(writer, sheet_name='Metadata', index=False)
                logger.info(f"✅ ذخیره شد: {excel_path}")
                # add a Header sheet with project info and filters summary, styled
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import Font

                    wb = load_workbook(excel_path)
                    # create header sheet at the front
                    if 'Header' in wb.sheetnames:
                        hdr = wb['Header']
                    else:
                        hdr = wb.create_sheet(title='Header', index=0)

                    hdr['A1'] = args.project_name
                    hdr['A2'] = f"Contact: {args.contact_email or ''}"
                    hdr['A3'] = f"Date: {date.today().isoformat()}"
                    hdr['A4'] = f"Rows: {len(top_20)}"
                    hdr['A1'].font = Font(size=14, bold=True)
                    hdr['A2'].font = Font(bold=False)

                    start_row = 6
                    hdr.cell(row=start_row, column=1, value='Filter Reason').font = Font(bold=True)
                    hdr.cell(row=start_row, column=2, value='Count').font = Font(bold=True)
                    r = start_row + 1
                    for reason, count in filter_obj.filters_applied.items():
                        hdr.cell(row=r, column=1, value=reason)
                        hdr.cell(row=r, column=2, value=count)
                        r += 1

                    wb.save(excel_path)
                except Exception as e:
                    logger.warning(f"⚠️ خطا هنگام افزودن شیت سربرگ به اکسل: {e}")
            except Exception as e:
                logger.warning(f"⚠️ خطا هنگام ذخیرهٔ اکسل: {e}")

        # ذخیره PDF در صورت درخواست و در دسترس بودن reportlab
        if args.save_pdf:
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont

                # Attempt to find a suitable TTF font for Persian/Arabic rendering
                def find_font_path():
                    candidates = [
                        os.path.join(os.getcwd(), 'Vazir.ttf'),
                        os.path.join(os.getcwd(), 'Vazirmatn.ttf'),
                        r'C:\Windows\Fonts\Tahoma.ttf',
                        r'C:\Windows\Fonts\Arial.ttf',
                        r'C:\Windows\Fonts\DejaVuSans.ttf',
                    ]
                    for p in candidates:
                        if p and os.path.exists(p):
                            return p
                    return None

                font_path = find_font_path()
                font_name = 'Helvetica'
                if font_path:
                    try:
                        font_name = 'CustomFont'
                        pdfmetrics.registerFont(TTFont(font_name, font_path))
                    except Exception:
                        font_name = 'Helvetica'

                pdf_name = f"top_symbols_{safe_proj}_{date.today().isoformat()}.pdf"
                pdf_path = os.path.join(out_dir, pdf_name)

                doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
                styles = getSampleStyleSheet()
                # override or create styles to use chosen font
                title_style = ParagraphStyle('Title', parent=styles['Title'], fontName=font_name, fontSize=16, leading=20)
                normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=font_name, fontSize=10, leading=12)
                heading_style = ParagraphStyle('Heading2', parent=styles.get('Heading2', styles['Normal']), fontName=font_name, fontSize=12, leading=14)

                elems = []

                # Header block
                header = f"{args.project_name}"
                elems.append(Paragraph(header, title_style))
                info_lines = [f"Date: {date.today().isoformat()}"]
                if args.contact_email:
                    info_lines.append(f"Contact: {args.contact_email}")
                for line in info_lines:
                    elems.append(Paragraph(line, normal_style))
                elems.append(Spacer(1, 12))

                elems.append(Paragraph(f"Top {len(top_20)} Symbols", heading_style))
                elems.append(Spacer(1, 6))

                # prepare table data (limited columns)
                cols = ['نماد', 'نام']
                if 'حجم' in top_20.columns:
                    cols.append('حجم')
                if 'آخرین معامله - مقدار' in top_20.columns:
                    cols.append('آخرین معامله - مقدار')

                table_data = [cols]
                for _, r in top_20[cols].iterrows():
                    row_vals = [str(r[c]) for c in cols]
                    table_data.append(row_vals)

                # calculate column widths roughly based on page width
                from reportlab.lib.pagesizes import A4 as A4_size
                page_w, page_h = A4_size
                usable_w = page_w - doc.leftMargin - doc.rightMargin
                # allocate more width to 'نام'
                ncols = len(cols)
                widths = []
                for c in cols:
                    if c == 'نام':
                        widths.append(usable_w * 0.5)
                    elif c == 'نماد':
                        widths.append(usable_w * 0.15)
                    else:
                        widths.append(usable_w * 0.35 / max(1, ncols-2))

                table = Table(table_data, repeatRows=1, hAlign='LEFT', colWidths=widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#d3d3d3')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('LEFTPADDING', (0,0), (-1,-1), 6),
                    ('RIGHTPADDING', (0,0), (-1,-1), 6),
                ]))
                elems.append(table)
                doc.build(elems)
                logger.info(f"✅ ذخیره شد: {pdf_path}")
            except ImportError:
                logger.warning('reportlab نصب نشده؛ برای تولید PDF، آن را نصب کنید: pip install reportlab')
            except Exception as e:
                logger.warning(f"⚠️ خطا هنگام تولید PDF: {e}")

    except Exception as e:
        logger.warning(f"⚠️ خطا هنگام ذخیره‌سازی جزئیات: {e}")
    
    # نمایش نتایج
    print("\n📍 برترین 20 نماد معتبر:\n")
    
    display_cols = ['نماد', 'نام']
    if 'حجم' in top_20.columns:
        display_cols.append('حجم')
    if 'آخرین معامله - مقدار' in top_20.columns:
        display_cols.append('آخرین معامله - مقدار')
    
    display_df = top_20[display_cols].copy()
    display_df.index = range(1, len(display_df) + 1)
    
    print(display_df.to_string())
    print("\n" + "="*80)
    print("✅ انجام شد!")
    print("="*80 + "\n")


if __name__ == "__main__":
    main()
